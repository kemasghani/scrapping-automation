from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")


def _fmt_sheet(ws, df: pd.DataFrame, *, freeze: str = "A2", zebra: bool = True) -> None:
    if df.empty:
        return
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
    ws.freeze_panes = freeze
    ws.row_dimensions[1].height = 28

    for col_idx, col in enumerate(df.columns, 1):
        letter = get_column_letter(col_idx)
        sample = [str(col)] + [str(v) for v in df[col].head(50).tolist() if v is not None]
        width = min(max(len(s) for s in sample) + 2, 55)
        ws.column_dimensions[letter].width = max(width, 12)

    wrap = Alignment(wrap_text=True, vertical="top")
    for row_idx in range(2, len(df) + 2):
        for cell in ws[row_idx]:
            cell.alignment = wrap
        if zebra and row_idx % 2 == 0:
            for cell in ws[row_idx]:
                if cell.fill.fgColor.rgb in (None, "00000000", "FFFFFFFF"):
                    cell.fill = ZEBRA_FILL


def write_workbook(
    path: Path,
    *,
    summary: pd.DataFrame,
    merchant: pd.DataFrame,
    menu: pd.DataFrame,
    modifiers: pd.DataFrame,
    promotions: pd.DataFrame,
    dietary: pd.DataFrame,
    field_guide: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        merchant.to_excel(writer, sheet_name="Merchant", index=False)
        menu.to_excel(writer, sheet_name="Menu", index=False)
        modifiers.to_excel(writer, sheet_name="Modifiers", index=False)
        promotions.to_excel(writer, sheet_name="Promotions", index=False)
        if not dietary.empty:
            dietary.to_excel(writer, sheet_name="Dietary", index=False)
        field_guide.to_excel(writer, sheet_name="Field Guide", index=False)

        _fmt_sheet(writer.sheets["Summary"], summary, zebra=False)
        _fmt_sheet(writer.sheets["Merchant"], merchant, zebra=False)
        _fmt_sheet(writer.sheets["Menu"], menu)
        _fmt_sheet(writer.sheets["Modifiers"], modifiers)
        _fmt_sheet(writer.sheets["Promotions"], promotions)
        if not dietary.empty:
            _fmt_sheet(writer.sheets["Dietary"], dietary)
        _fmt_sheet(writer.sheets["Field Guide"], field_guide, zebra=False)

        if not menu.empty:
            ws = writer.sheets["Menu"]
            ws.auto_filter.ref = ws.dimensions
            cols = list(menu.columns)

            def col_letter(name: str) -> str | None:
                return get_column_letter(cols.index(name) + 1) if name in cols else None

            avail_col = col_letter("available")
            promo_col = col_letter("promo_amount")
            dup_col = col_letter("is_duplicate_appearance")
            for row in range(2, len(menu) + 2):
                if avail_col:
                    c = ws[f"{avail_col}{row}"]
                    c.fill = GREEN if c.value in (True, "True") else RED
                if promo_col:
                    c = ws[f"{promo_col}{row}"]
                    if c.value not in (None, "", 0):
                        c.fill = YELLOW
                if dup_col:
                    c = ws[f"{dup_col}{row}"]
                    if c.value in (True, "True"):
                        c.font = Font(italic=True, color="808080")

            for money_col_name in (
                "price_before_promo",
                "price_after_promo",
                "promo_amount",
                "takeaway_price",
                "takeaway_discounted",
            ):
                letter = col_letter(money_col_name)
                if not letter:
                    continue
                for row in range(2, len(menu) + 2):
                    ws[f"{letter}{row}"].number_format = '"Rp" #,##0'
            pct_letter = col_letter("promo_percentage")
            if pct_letter:
                for row in range(2, len(menu) + 2):
                    ws[f"{pct_letter}{row}"].number_format = '0.0"%"'

        if not modifiers.empty and "price" in modifiers.columns:
            ws = writer.sheets["Modifiers"]
            ws.auto_filter.ref = ws.dimensions
            letter = get_column_letter(list(modifiers.columns).index("price") + 1)
            for row in range(2, len(modifiers) + 2):
                ws[f"{letter}{row}"].number_format = '"Rp" #,##0'
